# Libraries

import pandas as pd
import openpyxl


#Nome do jogador
name = str(input("Player name: ")).lstrip().rstrip()

#Dicionário com as posições
positions = {'CA': 'Centro-avante', 'SA': 'Segundo atacante', 'PD': 'Ponta direita', 'PE': 'Ponta esquerda', 'MC': 'Meia central', 'MEI': 'Meia', 
'MD': 'Meia direita', 'ME': 'Meia esquerda', 'VOL': 'Volante', 'ZGC': 'Zagueiro Central', 'ZGE': 'Zagueiro Esquerdo', 'ZGD': 'Zagueiro Direito', 'LD': 'Lateral Direito',
'LE': 'Lateral Esquerdo', 'GOL': 'Goleiro'}

#Adicionar posição
while True:
    for x in positions:
        print(f'{x} - {positions[x]}')
    position = str(input("Player Position: ")).strip().upper()
    
    player_positions = []
    print('\n')
    
    
    if position not in positions:
        print("Você não digitou uma das posições.\n")
        continue
    else:
        player_positions.append(position)
        print(f'Posição ({position}) adicionada.\n')
        
        break
    



#atributos
atributes = ["Attacking", "Technique", "Tactic", "Defense", "Creativity"]

#valores de cada atributo
values = []

print("Adicione os atributos do jogador.")
for atributo in atributes:
    value = int(input(atributo + ": "))
    values.append(value)

overall = int(sum(values)/len(values))

# Set data
df = pd.DataFrame({
f'Name': name,
f'Position': position,
f'ATT': [values[0]],
f'TEC': [values[1]],
f'TAC': [values[2]], 
f'DEF': [values[3]],
f'CRE': [values[4]],
f'OVERALL': overall
})




#excel path
xlsx = "players_stats\players_atributes.xlsx"

# Abre o arquivo Excel
wb = openpyxl.load_workbook(xlsx)
ws = wb.active

# Seleciona a planilha "players_atributes"

ultima_linha = ws.max_row+1
# Itera pelas linhas do DataFrame
for i, x in enumerate(df.values[0], 1):
    
    ws.cell(column=i, row = ultima_linha, value=x)
    
    
    

# Salva as alterações no arquivo Excel
wb.save(xlsx)